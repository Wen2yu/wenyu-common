# -*- coding:utf-8 -*-

# Name: excel_utils
# Product_name: PyCharm
# Description:
# Author : 'zhangjiawen'
# Date : 2021/8/6 10:29

import math
import os
import re
import sys
import traceback

from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell
from openpyxl.comments import Comment
from openpyxl.styles import Side, Border, Alignment, Font, PatternFill
from openpyxl.styles.colors import BLACK
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from jy_db import get_db, handle_param
from jy_utils.date_utils import day_str, datetime, last_month_begin, str_to_date, timedelta, ZERO_DATETIME
from jy_utils.dict_utils import deep_update, left_union
from jy_utils.http_utils import pool_req
from jy_utils.json_utils import dumps, load_json_file
from jy_utils.list_utils import first, to_list
from jy_utils.obj_utils import none
from jy_utils.str_utils import zip_str, str2dict
from jy_utils.sys_utils import str_val

base_path = 'D:\\excel\\'
# form_name = '全站注册会员分平台消费roi'
# result_file_name = '全站注册会员分平台消费roi(20.12)'
# form_name = '月全站收入完成情况(流水)'
# result_file_name = '2021.10月全站收入完成情况(流水)'

GET_DATA = 'get_data'

excel_0_date = datetime(1899, 12, 30)
cell_loc_pat = re.compile(r'([A-Z]+)+?([\+-]+)?(\d*)+?([\+-]+)?')


def cell_loc2idx(loc: str, n=0):
    """
    单元格位置字符串转坐标
    :param loc: 单元格位置字符串
    :param n: 偏移量
    :return:
    """
    col_name, col_extend, row, row_extend = cell_loc_pat.findall(loc)[0]
    return int(row) + eval(f'{row_extend}{n}'), column_index_from_string(col_name) + eval(f'{col_extend}{n}'), row_extend, col_extend


def cell_idx2loc(row, col, col_extend='', row_extend=''):
    """
    单元格坐标转位置字符串
    :param row: 行号
    :param col: 列号
    :param col_extend: 列扩展方向
    :param row_extend: 行扩展方向
    :return:
    """
    # return f'{"".join([chr(i + 64) for i in int2list(col, divide_num=26)])}{col_extend}{row}{row_extend}'
    return f'{get_column_letter(col)}{col_extend}{row}{row_extend}'


def cell_loc_generator(loc: str):
    """

    :param loc:
    :return:
    """
    row, col, row_extend, col_extend = cell_loc2idx(loc)
    idx = 0
    while 1:
        if row_extend:
            yield cell_idx2loc(row, col)
            if col_extend:
                col += eval(f'{col_extend}1')
            row += eval(f'{row_extend}1')
        elif col_extend:
            yield cell_idx2loc(row, col)
            col += eval(f'{col_extend}1')
        else:
            yield loc
            return
        idx += 1
        if idx >= 100000:
            break


def cell_loc_range2locs(locs_str: str):
    """
    单元格位置区间转单元格位置字符串列表
    :param locs_str: 单元格位置区间字符串
    :return:
    """
    result = ''
    for locs_str in locs_str.split(','):
        if locs_str.find(':') > 0:
            b_row, b_col, e_row, e_col, row_extend, col_extend, loc_list = 0, 0, 0, 0, '', '', []
            b_e = locs_str.split(':')
            b_loc = b_e[0]
            if b_loc:
                b_row, b_col, row_extend, col_extend = cell_loc2idx(b_loc)
            if len(b_e) > 1:
                e_loc = b_e[-1]
                if e_loc:
                    e_row, e_col = cell_loc2idx(e_loc)[:2]
            for i in range(b_row, (e_row or b_row) + 1):
                for j in range(b_col, (e_col or b_col) + 1):
                    loc_list.append(cell_idx2loc(i, j, col_extend, row_extend))
            result += f',{",".join(loc_list)}'
        else:
            result += f',{locs_str}'
    return result[1:]


def datetime2exceldate(d: datetime):
    if type(d) != datetime:
        return d
    delt = d - excel_0_date
    return delt.days - (delt.days < 60) + (delt.seconds / 86400 if delt.seconds else 0)


def _cell_value(v):
    """

    :param v:
    :return:
    """
    if v == '-':
        v = None
    return datetime2exceldate(cell_value2datetime(v)) if v and (
            type(v) == datetime or (type(v) == str and len(v) >= 4)
    ) else v


def _cell_type_value(tpe, v):
    """

    :param type:
    :param v:
    :return:
    """
    if tpe in (int, float):
        v = _cell_value(v)
    elif tpe == str:
        return day_str(v) if type(v) == datetime else str(v)
    elif tpe == type(None) :
        v = _cell_value(v) if type(v) == str else v
    return v


def _cell_data(cell, memo=False):
    if memo:
        return cell.comment.text if cell.comment else ''
    else:
        return _cell_value(cell.value)


def read_cell_data(sheet: Worksheet, loc: str, memo=False):
    """
    读取单元格值
    :param sheet:
    :param loc:
    :param memo: False，获取单元格值；True，获取单元格批注
    :return:
    """
    value_dict = dict()
    row, col, row_extend, col_extend = cell_loc2idx(loc)
    cell = sheet.cell(row, col)
    if row_extend:
        while cell.value:
            value_dict[cell_idx2loc(row, col)] = _cell_data(cell, memo=memo)
            if col_extend:
                col += 1
                cell = sheet.cell(row, col)
                while cell.value:
                    value_dict[cell_idx2loc(row, col)] = _cell_data(cell, memo=memo)
            row += 1
            cell = sheet.cell(row, col)
    elif col_extend:
        while cell.value:
            value_dict[cell_idx2loc(row, col)] = _cell_data(cell, memo=memo)
            col += 1
            cell = sheet.cell(row, col)
    else:
        value_dict[cell_idx2loc(row, col)] = _cell_data(cell, memo=memo)
    return value_dict


def read_history_data(formwork: Workbook, data_conf: dict):
    """
    从模板中读取历史数据
    :param formwork:
    :param data_conf:
    :return:
    """
    def_conf = data_conf.get('default', dict())
    history_data = dict()
    for sheet_name in formwork.sheetnames:
        sheet: Worksheet = formwork[sheet_name]
        history_data[sheet_name] = dict()
        get_data_conf = data_conf.get(sheet_name, dict()).get('get_data', dict()) or def_conf.get('get_data', dict())
        data_locs = ','.join([cell_loc_range2locs(key) for key in get_data_conf.keys()]).split(',')
        history_data[sheet_name] = dict()
        for data_loc in data_locs:
            if data_loc:
                history_data[sheet_name].update(read_cell_data(sheet, data_loc))
        get_memo_conf = data_conf.get(sheet_name, dict()).get('get_memo', dict()) or def_conf.get('get_memo', dict())
        memo_locs = ','.join([cell_loc_range2locs(key) for key in get_memo_conf.keys()]).split(',')
        history_data[sheet_name]['memo'] = dict()
        for memo_loc in memo_locs:
            if memo_loc:
                history_data[sheet_name]['memo'].update(read_cell_data(sheet, memo_loc, memo=True))
    return history_data


def extract_formwork(formname, group='reports', excel_type='xlsx', data_conf_sheet_name='数据配置', **kw):
    """
    提取模板excel历史数据及配置
    :param formname: 模板名称
    :param group: 模板分组
    :param excel_type: 模板文件类型
    :param data_conf_sheet_name: 数据配置sheet_name
    :return:
    """
    path = f'{base_path}{group}\\{formname}.{excel_type}'
    formwork: Workbook = load_workbook(path)
    data_conf = dict()
    if data_conf_sheet_name:
        data_conf_sheet: Worksheet = formwork[data_conf_sheet_name]
        row, column, sheet_name = 2, 1, ''
        cell = data_conf_sheet.cell(row, column) if data_conf_sheet else None
        while (cell and cell.value) or sheet_name:
            while cell.value or sheet_name:
                if cell.value and cell.column == 1:
                    sheet_name = cell.value
                    data_conf[sheet_name] = dict()
                elif cell.column == 2:
                    k = cell.value
                    if not k:
                        sheet_name = ''
                        break
                elif cell.column == 3:
                    if cell.value == '-':
                        data_conf[sheet_name][k] = dict()
                        column += 1
                        cell = data_conf_sheet.cell(row, column)
                        data_k = cell.value
                        data_conf[sheet_name][k][data_k] = list()
                        while cell.value:
                            column += 1
                            cell = data_conf_sheet.cell(row, column)
                            if cell.value:
                                data_conf[sheet_name][k][data_k].append(cell.value)
                    elif cell.value == '|':
                        data_conf[sheet_name][k], data_k = dict(), ''
                        while cell.value or data_k:
                            row += 1
                            column = 3
                            cell = data_conf_sheet.cell(row, column)
                            if cell.value and cell.value != '-':
                                data_k = cell.value
                            else:
                                break
                            column += 1
                            cell = data_conf_sheet.cell(row, column)
                            data_conf[sheet_name][k][data_k] = cell.value
                    else:
                        data_conf[sheet_name][k] = cell.value
                if sheet_name and column < 3:
                    column += 1
                else:
                    break
                cell = data_conf_sheet.cell(row, column)
            row, column = row + 1, 1
            cell = data_conf_sheet.cell(row, column)
        formwork.remove(data_conf_sheet)
    if not data_conf:
        data_conf = load_json_file(path.replace(f'.{excel_type}', '_data_conf.json'))
    if kw.get('conf'):
        data_conf = deep_update(data_conf, kw['conf'])
    with open(path.replace(f'.{excel_type}', '_data_conf.json'), 'w') as f:
        f.write(dumps(data_conf, indent=2))
    history_data = read_history_data(formwork, data_conf)
    with open(path.replace(f'.{excel_type}', '_history_data.json'), 'w') as f:
        f.write(dumps(history_data, indent=2))
    formwork.save(path.replace(f'.{excel_type}', f'_form.{excel_type}'))


def copy_sheet(src_sheet: Worksheet, aim_sheet: Worksheet, remain_rows=0, remain_cols=0):
    """
    带格式复制sheet
    :param src_sheet: 源sheet
    :param aim_sheet: 目标sheet
    :param remain_rows: 保留行数，<= 0:全保留
    :param remain_cols: 保留列数，<= 0:全保留
    :return:
    """
    # tab color
    aim_sheet.sheet_properties.tabColor = src_sheet.sheet_properties.tabColor
    # 开始处理合并单元格形式为“(<CellRange A1：A4>,)，替换掉(<CellRange 和 >,)' 找到合并单元格
    wmcs = list(src_sheet.merged_cells)
    if len(wmcs) > 0:
        for wmc in wmcs:
            aim_cell_range = str(wmc).replace('(<CellRange ', '').replace('>,)', '')
            aim_sheet.merge_cells(aim_cell_range)
    row_idx = 0
    for i, row in enumerate(src_sheet.iter_rows()):
        aim_sheet.row_dimensions[i + 1].height = src_sheet.row_dimensions[i + 1].height
        col_idx = 0
        for j, cell in enumerate(row):
            col_letter = get_column_letter(j + 1)
            aim_sheet.column_dimensions[col_letter].width = src_sheet.column_dimensions[col_letter].width
            aim_sheet.cell(row=i + 1, column=j + 1, value=cell.value)
            # 设置单元格格式
            copy_cell_style(src_sheet.cell(i + 1, j + 1), aim_sheet.cell(i + 1, j + 1))
            col_idx += 1
            if 0 < remain_cols <= col_idx:
                break
        row_idx += 1
        if 0 < remain_rows <= row_idx:
            break


def copy_cell_style(src_cell, aim_cell):
    """
    复制单元格格式
    :param src_cell:
    :param aim_cell:
    :return:
    """
    if src_cell.has_style:
        try:
            aim_cell.fill = copy(src_cell.fill)
        except Exception:
            traceback.print_exc()
        aim_cell.font = copy(src_cell.font)
        aim_cell.border = copy(src_cell.border)
        aim_cell.protection = copy(src_cell.protection)
        aim_cell.alignment = copy(src_cell.alignment)
        aim_cell.number_format = copy(src_cell.number_format)
        aim_cell._style = copy(src_cell._style)
        pass
    if src_cell.comment:
        aim_cell.comment = Comment('memo:', 'memo', src_cell.comment.height, src_cell.comment.width)
    if not (src_cell.value in [None, '-', ''] or aim_cell.value in [None, '-', '']) and not (
        type(src_cell.value) in (int, float)
    ) and type(aim_cell.value) != str and type(src_cell.value) != type(aim_cell.value):
        print(src_cell, type(src_cell.value), src_cell.value, src_cell.number_format, aim_cell, type(aim_cell.value), aim_cell.value, aim_cell.number_format)
        src_cell.value = _cell_type_value(type(aim_cell.value), src_cell.value)
        print(src_cell, type(src_cell.value), src_cell.value, src_cell.number_format, aim_cell, type(aim_cell.value), aim_cell.value, aim_cell.number_format)


def write_sheet_data(sheet: Worksheet, data: dict):
    """
    向目标sheet写入数据
    :param sheet: 目标sheet
    :param data: 待写入数据
    :return:
    """
    for k, v in data.items():
        if k == 'memo':
            for loc, memo in data['memo'].items():
                cell = sheet.cell(*cell_loc2idx(loc)[:2])
                if cell.comment:
                    cell.comment.text = memo
                else:
                    cell.comment = Comment(memo, 'memo')
        else:
            sheet.cell(*cell_loc2idx(k)[:2], v[0] if type(v) in (tuple, list) else v)


def handle_data_conf(data_conf: str, sp_chars='://|', **kw):
    """
    处理数据配置，获取数据
    :param data_conf:
    :param sp_chars:
    :return:
    """
    ks, d = data_conf.split(sp_chars), []
    if len(ks) > 1:
        conf_type = ks[0]
        if conf_type.startswith("sql"):
            db, sql_params = get_db(*tuple(ks[1].split(':'))), dict()
            if len(ks) > 3:
                sql_params = str2dict(ks[3], ['&', '='])
            for k, v in sql_params.items():
                if type(v) == str and v.startswith('#') and kw.get(v[1:]):
                    sql_params[k] = kw[v[1:]]
            sql_params.update(kw)
            sql_params['include_col_names'] = sql_params.get('include_col_names', 1)
            return db.select(sql=ks[2], **sql_params) if db and len(ks) > 2 else []
        elif conf_type.startswith('param'):
            return [[handle_param(kw.get(param_name, ''), **kw) for param_name in ks[1].split(',')]]
        elif conf_type.startswith('formula'):
            return [[ks[1]]]
        elif conf_type.startswith('http'):
            if len(ks) > 2:
                return [[pool_req(ks[2], method=ks[1] if ks[1] else None, params=ks[3] if len(ks) > 3 else None)]]
        elif conf_type.startswith('list'):
            s = str_val(ks[1])
            print(s, type(s))
            return [ks[1].split(',') if type(s) == str else to_list(s)]
    return d


def cell_value2datetime(cell_value, fmt=''):
    """

    :param fmt:
    :param cell_value:
    :return:
    """
    if type(cell_value) in (int, float) and 0 < cell_value < 109575 and cell_value != 60:
        return excel_0_date + timedelta(days=cell_value + int(cell_value < 60))
    elif type(cell_value) == str:
        try:
            return str_to_date(cell_value, fmt=fmt)
        except Exception as e:
            return str_val(cell_value)
    else:
        return cell_value


def cell_value_equals(cell_value1, cell_value2):
    """
    判断单元格值是否相等
    :param cell_value1:
    :param cell_value2:
    :return:
    """
    if cell_value1 is None:
        return False
    elif type(cell_value2) == type(cell_value1) == str or (
            type(cell_value1) in (int, float, str, datetime) and type(cell_value2) in (datetime, str)
    ):
        try:
            return cell_value2datetime(cell_value1) == cell_value2datetime(cell_value2) or \
                   cell_value1 == type(cell_value1)(cell_value2)
        except Exception:
            return cell_value1 == cell_value2
    elif type(cell_value1) in (int, float):
        return cell_value2 and cell_value1 == type(cell_value1)(cell_value2)
    else:
        return cell_value1 == cell_value2


def zip_cell_locs_data(sheetname, history_data, cell_locs: str, d: list, batch_key_cell=None, memo=False, **kw):
    """

    :param sheetname:
    :param history_data:
    :param cell_locs:
    :param d:
    :param batch_key_cell:
    :param memo: 增加批注
    :return:
    """
    result = dict()
    if not history_data.get(sheetname):
        history_data[sheetname] = dict()
    if memo and not history_data[sheetname].get('memo'):
        history_data[sheetname][memo] = dict()
    cell_locs = cell_loc_range2locs(cell_locs).split(',')
    size = len(cell_locs)
    if d:
        d = d * math.ceil(size / len(d))
    batch_key_row, batch_key_col = 0, 0
    for i in range(size):
        cell_loc = cell_locs[i]
        data = none(cell_value2datetime(d[i]) if type(d[i]) == str and len(d[i]) in [10, 19] else str_val(d[i]), '')
        if type(data) == datetime:
            data = datetime2exceldate(data)
        row, col, row_extend, col_extend = cell_loc2idx(cell_loc)
        style_extend = 0
        s_row, s_col = row, col
        if not (batch_key_row or batch_key_col) and (batch_key_cell or row_extend or col_extend):
            batch_key_row, batch_key_col = row, col
            batch_key_cell_loc_gen = cell_loc_generator(batch_key_cell if batch_key_cell else cell_locs[0])
            while 1:
                try:
                    batch_key_cell_loc = next(batch_key_cell_loc_gen)
                    if batch_key_cell:
                        if history_data[sheetname].get(batch_key_cell_loc) is None or cell_value_equals(
                            to_list(history_data[sheetname].get(batch_key_cell_loc))[0],
                            kw.get('batch_key_value', handle_param(kw.get('tdate', day_str(last_month_begin())), **kw))
                        ):
                            batch_key_row, batch_key_col = cell_loc2idx(batch_key_cell_loc)[:2]
                            break
                        else:
                            continue
                    else:
                        if history_data[sheetname].get(batch_key_cell_loc) is None:
                            batch_key_row, batch_key_col = cell_loc2idx(batch_key_cell_loc)[:2]
                            break
                except StopIteration:
                    break
        if row_extend:
            row, style_extend = batch_key_row, int(s_row != batch_key_row)
            # s_row = row - 1
        elif col_extend:
            col, style_extend = batch_key_col, 2 if s_col != batch_key_col else 0
            # s_col = col - 1
        print(f'++++++++_______{history_data[sheetname].get(cell_idx2loc(s_row, s_col)), type(history_data[sheetname].get(cell_idx2loc(s_row, s_col))), data}')
        result[cell_idx2loc(row, col)] = (_cell_type_value(
            type(history_data[sheetname].get(cell_idx2loc(s_row, s_col))), data.replace(
                '#col_name', get_column_letter(col)
            ).replace('#row', str(row)).replace(
                '#col', str(col)
            ) if type(data) == str and data and data[0] == '=' else data
        ), style_extend)
    return {sheetname: result} if not memo else {sheetname: {
        'memo': {k: v[0] if type(v) == tuple else v for k, v in result.items()}}
    }


def add_data2history_data(book: Workbook, sheet_name, data_conf: dict, history_data: dict, cell_locs: str, d: list,
                          batch_key_cell=None, memo=False, **kw):
    """
    增加数据到history_data
    :param book:
    :param sheet_name:
    :param data_conf:
    :param history_data:
    :param cell_locs:
    :param d:
    :param batch_key_cell:
    :param memo: 增加批注
    :return:
    """
    if type(d) == tuple:
        col_names, d = [i[0] for i in d[1]], d[0]
    else:
        col_names = []
    if sheet_name == 'default':
        sheet_name_idx = None
        for i in range(len(col_names)):
            if col_names[i].upper() == 'SHEET_NAME':
                sheet_name_idx = i
                col_names.pop(i)
                break
        if sheet_name_idx is not None:
            sheet_names = [name for name in book.sheetnames if name not in data_conf.keys()]
            last_sheet_name = sheet_names[-1] if sheet_names else None
            for dd in d:
                dd = list(dd)
                sheet_name = dd.pop(sheet_name_idx)
                if not history_data.get(sheet_name):
                    history_data[sheet_name] = dict()
                    if sheet_name in book.sheetnames:
                        last_sheet_name = sheet_name
                    else:
                        book.create_sheet(sheet_name)
                    if last_sheet_name:
                        copy_sheet(
                            book[last_sheet_name], book[sheet_name],
                            remain_rows=data_conf.get('default', dict()).get('remain_rows', 0),
                            remain_cols=data_conf.get('default', dict()).get('remain_cols', 0)
                        )
                    last_sheet_name = sheet_name
                history_data = deep_update(
                    history_data, zip_cell_locs_data(sheet_name, history_data, cell_locs, dd, batch_key_cell, memo=memo, **kw),
                    merge_list=False, merge_obj2list=False
                )
        else:
            for k in history_data:
                if not data_conf.get(k):
                    for dd in d:
                        history_data = deep_update(
                            history_data, zip_cell_locs_data(k, history_data, cell_locs, to_list(dd), batch_key_cell, memo=memo, **kw),
                            merge_list=False, merge_obj2list=False
                        )
    else:
        idx = 0
        for dd in d:
            idx += 1
            history_data = deep_update(
                history_data, zip_cell_locs_data(sheet_name, history_data, cell_locs, to_list(dd), batch_key_cell, memo=memo, **kw),
                merge_list=False, merge_obj2list=False
            )


def write_formwork(formname, file_name, group='reports', excel_type='xlsx', **kw):
    """

    :param formname:
    :param file_name:
    :param group:
    :param excel_type:
    :param kw:
    :return:
    """
    if not os.path.exists(f'{base_path}{group}'):
        os.mkdir(f'{base_path}{group}')
    if not os.path.exists(f'{base_path}{group}/form_result'):
        os.mkdir(f'{base_path}{group}/form_result')
    base_form_path = f'{base_path}{group}/{formname}_form.{excel_type}'
    # base_formwork: Workbook = load_workbook(base_form_path)
    # aim_form_work = Workbook()
    # for sheet_name in base_formwork.sheetnames:
    #     copy_sheet(base_formwork[sheet_name], aim_form_work.create_sheet(sheet_name))
    new = 0
    aim_formwork: Workbook
    try:
        aim_formwork = load_workbook(base_form_path)
    except Exception:
        new = 1
        aim_formwork = Workbook()
    data_conf: dict = deep_update(
        none(load_json_file(base_form_path.replace(f'_form.{excel_type}', '_data_conf.json')), dict()),
        kw.pop('data_conf', dict())
    )
    history_data = load_json_file(base_form_path.replace(f'_form.{excel_type}', '_history_data.json'))
    for k in data_conf:
        if k != 'default' and not history_data.get(k):
            history_data[k] = dict()
    if kw.get('save_data_conf'):
        with open(f'{base_path}{group}/{formname}_data_conf.json', 'w') as f:
            f.write(dumps(data_conf, indent=2))
        with open(f'{base_path}{group}/{formname}_history_data.json', 'w') as f:
            f.write(dumps(history_data, indent=2))
    last_sheet_name = aim_formwork.sheetnames[0]
    if not history_data:
        history_data[last_sheet_name] = dict()
    for sheet_name, data in history_data.items():
        if not aim_formwork.__contains__(sheet_name):
            copy_sheet(
                aim_formwork[last_sheet_name],
                aim_formwork.create_sheet(sheet_name),
                remain_rows=data_conf.get('default', dict()).get('remain_rows', 0),
                remain_cols=data_conf.get('default', dict()).get('remain_cols', 0)
            )
        # write_sheet_data(aim_formwork[sheet_name], data)
        last_sheet_name = sheet_name
    def_conf: dict = dict()
    merge_conf: dict = dict()
    for sheet_name, sheet_conf in data_conf.items():
        if sheet_name == 'default':
            def_conf = sheet_conf
        else:
            sheet_conf = left_union(sheet_conf, def_conf)
        if sheet_conf.get('merge_cells'):
            merge_conf[sheet_name] = {'merge_cells': sheet_conf['merge_cells'], 'col_first': sheet_conf.get('col_first', 0)}
        for cell_locs, conf in sheet_conf.get('get_data', dict()).items():
            add_data2history_data(
                aim_formwork, sheet_name, data_conf, history_data, cell_locs, handle_data_conf(conf, **kw),
                batch_key_cell=sheet_conf.get('batch_key_cell'), **kw
            )
        for cell_locs, conf in sheet_conf.get('get_memo', dict()).items():
            add_data2history_data(
                aim_formwork, sheet_name, data_conf, history_data, cell_locs, handle_data_conf(conf, **kw),
                batch_key_cell=sheet_conf.get('batch_key_cell'), memo=True, **kw
            )
    for sheet_name, his_data in history_data.items():
        sheet: Worksheet = aim_formwork[sheet_name]
        if not his_data and kw.get('rmv_empty_data_sheet', 1):
            aim_formwork.remove(aim_formwork[sheet_name])
            continue
        if kw.get('hashed_password'):
            sheet.protection.sheet = True
            sheet.protection.enable()
            sheet.protection.password = kw['hashed_password']
        memo = his_data.get('memo', dict())
        _data_type = data_conf.get(sheet_name, data_conf.get('default', dict())).get('data_type', dict())
        data_type = dict()
        for k in _data_type:
            for loc in k.split(','):
                data_type[loc] = _data_type[k]
        header_line = data_conf.get(sheet_name, data_conf.get('default', dict())).get('header_line')
        for k, v in his_data.items():
            if k == 'memo':
                continue
            row, col = cell_loc2idx(k)[:2]
            style_extend = 0
            if type(v) in (tuple, list):
                style_extend, v = v[1], v[0]
            empty_replace = None
            if data_conf.get(sheet_name):
                empty_replace = data_conf[sheet_name].get('empty_replace')
            elif data_conf.get('default'):
                empty_replace = data_conf['default'].get('empty_replace')
            v = empty_replace if (none(v, '') == '' or v == 0) and empty_replace else none(v, '')
            try:
                cell: Cell = sheet.cell(row, col, value=(v))
            except Exception as e:
                traceback.print_exc()
                print(v, e)
            if style_extend:
                if style_extend == 1:
                    sheet.row_dimensions[row].height = sheet.row_dimensions[row - 1].height
                elif style_extend == 2:
                    col_letter = get_column_letter(col)
                    sheet.column_dimensions[col_letter].width = sheet.column_dimensions[get_column_letter(col-1)].width
                src_cell = sheet.cell(row - (style_extend == 1), col - (style_extend == 2))
                if src_cell.has_style:
                    copy_cell_style(src_cell, cell)
                else:
                    style_extend = 0
            else:
                if type(cell.value) == datetime and new:
                    cell._style = None
            if not (cell.has_style or style_extend):
                if data_type.get(cell.coordinate) == 'datetime':
                    cell.value = cell_value2datetime(cell.value, fmt='%Y-%m-%d %H:%M:%S')
                    cell.number_format = 'yyyy-mm-dd hh:mm:ss'
                elif data_type.get(cell.coordinate) == 'date':
                    cell.value = cell_value2datetime(cell.value, fmt='%Y-%m-%d')
                    cell.number_format = 'yyyy-mm-dd'
                elif type(cell.value) == datetime:
                    if cell.value.time() > ZERO_DATETIME.time():
                        cell.number_format = 'yyyy-mm-dd hh:mm:ss'
                    else:
                        cell.number_format = 'yyyy-mm-dd'
                col_letter = get_column_letter(col)
                thin = Side(border_style="thin", color=BLACK)
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                alignment = Alignment(horizontal='center', vertical='center')
                cell.alignment = alignment
                font = Font(name='宋体', sz=9, b=True)
                sheet.column_dimensions[col_letter].width = 20.88 if cell.number_format == 'yyyy-mm-dd hh:mm:ss' else (
                    10.88 if cell.number_format == 'yyyy-mm-dd' else max(
                        len(str(cell.value)) * 1.2 if type(cell.value) == str else 12,
                        sheet.column_dimensions[col_letter].width
                    )
                )
                if cell.row <= header_line:
                    _fill = PatternFill(start_color="AACF91", end_color="AACF91", fill_type="solid")
                    cell.fill = _fill
            if memo.get(k):
                if cell.comment:
                    cell.comment.text = memo[k]
                else:
                    cell.comment = Comment(memo[k], 'memo', min(len(memo[k]) * 2, 480), len(memo[k]) * 8 / min(len(memo[k]) * 2, 480) + 20)
        if sheet_name != 'Sheet' and (merge_conf.get(sheet_name) or merge_conf.get('default')):
            merge_conf = merge_conf.get(sheet_name, merge_conf.get('default'))
            merge_cells = merge_conf['merge_cells'].split(',')
            for merge_cell in merge_cells:
                merge_cell_by_str(sheet, merge_cell, col_first=merge_conf['col_first'])
    if len(aim_formwork.sheetnames) > 1 and 'Sheet' in aim_formwork.sheetnames:
        aim_formwork.remove(aim_formwork['Sheet'])
    for name in data_conf.get('default', dict()).get('remove_sheets', '').split(','):
        if name in aim_formwork.sheetnames:
            aim_formwork.remove(aim_formwork[name])
    if kw.get('hashed_password'):
        aim_formwork.security.set_workbook_password(kw['hashed_password'], already_hashed=True)
        aim_formwork.security.lockStructure = True
    aim_formwork.save(filename=f'{base_path}/{group}/form_result/{file_name}.{excel_type}')
    aim_formwork.close()
    history_data = {
        sheet_name: his_data for sheet_name, his_data in history_data.items() if data_conf.get(sheet_name, data_conf.get('default', dict())).get('batch_key_cell')
    }
    with open(base_form_path.replace(f'_form.{excel_type}', '_history_data.json'), 'w') as f:
        f.write(dumps(history_data, indent=2))


def merge_cell_by_str(sheet: Worksheet, merge_cell: str, col_first=0):
    """
    合并merge_cell范围内值相同的单元格
    :param sheet:
    :param merge_cell:
    :param col_first:
    :return:
    """
    b, e = tuple(merge_cell.split(':'))
    b_row, b_col = cell_loc2idx(b)[:2]
    if b and e:
        e_row, e_col = 9999, 9999
        row_extend, col_extend = int(e[-1] == '+'), int(e[0] == '+')
        if not row_extend and not col_extend:
            e_row, e_col = cell_loc2idx(e)[:2]
        elif row_extend and not col_extend:
            e_col = column_index_from_string(e[:-1])
        elif not row_extend and col_extend:
            e_row = int(e[1:])
        val = sheet[b].value
        while val is not None:
            row, col = b_row, b_col
            if e_row > b_row and e_col <= b_col:
                while val == sheet.cell(row + 1, col).value in [val, '', '-'] and row <= e_row:
                    row += 1
            elif e_col > b_col and e_row <= b_row:
                while val == sheet.cell(row, col + 1).value in [val, '', '-'] and col <= e_col:
                    col += 1
            elif e_row > b_row and e_col > b_col:
                _row, _col, _e_row, _e_col = b_row, b_col, 0, 0
                if col_first:
                    while sheet.cell(_row + 1, col).value in [val, '', '-'] and _row <= (_e_row or e_row):
                        _row += 1
                    _e_row = _row
                else:
                    while sheet.cell(row, _col + 1).value in [val, '', '-'] and _col <= (_e_col or e_col):
                        _col += 1
                    _e_col = _col
                if col_first:
                    while sheet.cell(row, col + 1).value in [val, '', '-'] and col <= e_col:
                        col += 1
                        _row = b_row
                        while sheet.cell(_row + 1, col).value in [val, '', '-'] and _row <= (_e_row or e_row):
                            _row += 1
                        _e_row = _row
                    row = _e_row
                else:
                    while sheet.cell(row + 1, col).value in [val, '', '-'] and row <= e_row:
                        row += 1
                        _col = b_col
                        while sheet.cell(row, _col + 1).value in [val, '', '-'] and _col <= (_e_col or e_col):
                            _col += 1
                        _e_col = _col
                    col = _e_col
            print('merge_cells(%d, %d, %d, %d)' % (b_row, b_col, row, col))
            sheet.merge_cells(None, b_row, b_col, min(row, e_row), min(col, e_col))
            if b_row < row < e_row or b_col < col < e_col:
                b_row, b_col = row + int(e_row > b_row), b_col if e_row > b_row else b_col + int(e_col > b_col)
                print(b_row, b_col, e_row, e_col)
                val = sheet.cell(b_row, b_col).value
            else:
                break

# extract_formwork(form_name)
# write_formwork(form_name, result_file_name, batch_key_value=datetime(2021, 10, 1), tdate='datestr#last_month_begin?fmt=%Y-%m-%d')
# write_formwork(form_name, result_file_name, tdate='datestr#last_month_begin?fmt=%Y-%m-%d')
# write_formwork(form_name, result_file_name, mth='datestr#last_month_begin?fmt=%Y%m', tdate='datestr#last_month_begin?fmt=%Y-%m-%d', end='datestr#month_begin?fmt=%Y-%m-%d')
extract_formwork(formname='Love21 Monthly Report Table(NEW)', data_conf_sheet_name='')
write_formwork('Love21 Monthly Report Table(NEW)', file_name='Love21 Monthly Report Table2021年12月(NEW)', tdate='datestr#last_month_begin?fmt=%Y-%m-%d')
