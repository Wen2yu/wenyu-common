# -*- coding: UTF8 -*-
'''
  导出对账结果excel
'''
# python37 /u01/vip48dian/lit/fentan/duizh/bin/exp_duizh_result.py 2022-08-01 liuquanjun1@jiayuan.com
# moudle
import cx_Oracle
import sys
from sys import path
import time
import json
import pandas as pd
import os
import numpy as np
import matplotlib.pyplot as plt
import datetime
from dateutil.relativedelta import relativedelta
from dbbox.dbmirror import DbMirror
from dbbox.dbsuite import db_common
from dbbox.dbshuffle import generate_insert
from jy_db import get_db, handle_param, parse_data
from jy_utils.date_utils import day_str, cur_day_begin, month_begin, add_days, last_month_begin
from jy_utils.operation_utils import add
import zipfile
import shutil
import openpyxl
import traceback


b_time = time.time()

# 传入日期
v_date = sys.argv[1]
pos_name = sys.argv[2]

# 导入方法
path.append('/u01/job/func')
from cx_Oracle_conn import __cx_Oracle_conn__
from pymysql_conn import __pymysql_conn__
from add_format import __add_format__
from write_sheet import __write_sheet__
from tup_to_dict import __tup_to_dict__

lbday_m = v_date[0: 7]
# jy_db = __cx_Oracle_conn__('newdatamart1')
# bh_db = cx_Oracle.connect('bh_finance/bh123456@newdatamart1')
rordb = get_db('Mysql', 'jydata_ror', auto_close=False)

sql_dict = {
  'collect_add': {
    'memo': '汇总新增表明细',
    'sql': '''
      select tdate     as "统计时间",
             orderno    as "订单号/收入类别",
             '0' as "产品id",
             type as "订单类别",
             uname      as "姓名",
             nc_shop_name      as "店铺",
             paytime   as "收款时间",
             cost       as "收款金额",
             balance    as "上期结余/本期新增",
             bdate      as "所属期间",
             valid      as "服务期间",
             '' as "收费标准/月",
             '' as "本期确认收入金额",
             '' as "VIP收入",
             '' as "线上资源使用收入",
             '' as "本期结余",
             '' as "税金",
             '' as "加盟本月分摊成本",
             0 as "分成率",
             '' as "负责人",
             crank      as "联营直营标识",
             ptype as "缘分期标识",
             itype   as "线上资源使用收入标识",
             '' as "订金合同编号",
             '' as "求婚宝合同编号",
             cust_id    as "资源ID",
             zt         as "账套",
             sup_id     as "补充协议ID",
             con_type   as "合同类型",
             con_money  as "合同金额",
             deduct_money  as "代金券抵扣金额",
             signed_type as "签单类型",
             service_dept as "现服务部门",
             lakala     as "拉卡拉",
             wangfujing as "数字王府井",
             remit_amt  as "银行到款",
             loan_amt   as "缘分期到款",
             remark as "备注",
             shop_id as "店铺ID",
             brand as "品牌"
        from cv_collect_result
       where tdate >= '2022-06-01' and nc_shop_name = '{sname}' 
       order by tdate, paytime, type, orderno
    '''
  },
  'lakala_detail': {
    'memo': 'pos到款拉卡拉明细',
    'sql': '''
      select tdate as "归属月",
             orderno as "合同编号",
             txdate as "刷卡日期",
             uname as "刷卡人",
             pos_amount as "金额",
             '' as "用途(VIP或地面活动)",
             terminalid as "刷卡机终端号",
             rrn as "交易参考号"
        from cv_pos_pay_order
       where tdate >= '2022-06-01' 
         and channel_type = 30 and nc_shop_name = '{sname}'
       order by tdate, txdate
    '''
  },
  'wangfujing_detail': {
    'memo': 'pos到款王府井明细',
    'sql': '''
      select tdate as "归属月",
             orderno as "合同编号",
             txdate as "刷卡日期",
             uname as "刷卡人",
             pos_amount as "金额",
             '' as "用途(VIP或地面活动)",
             terminalid as "刷卡机终端号",
             rrn as "交易参考号"
        from cv_pos_pay_order
       where tdate >= '2022-06-01' 
         and channel_type = 20 and nc_shop_name = '{sname}'
       order by tdate, txdate
    '''
  },
  'collect_update': {
    'memo': '汇总调账表明细',
    'sql': '''
      select tdate     as "统计时间",
             orderno    as "订单号/收入类别",
             pro_id as "产品id",
             type as "订单类别",
             uname      as "姓名",
             nc_shop_name      as "店铺",
             paytime   as "收款时间",
             cost       as "收款金额",
             balance    as "上期结余/本期新增",
             bdate      as "所属期间",
             valid      as "服务期间",
             '' as "收费标准/月",
             '' as "本期确认收入金额",
             '' as "VIP收入",
             '' as "线上资源使用收入",
             '' as "本期结余",
             '' as "税金",
             '' as "加盟本月分摊成本",
             srate as "分成率",
             '' as "负责人",
             crank      as "联营直营标识",
             ptype as "缘分期标识",
             itype   as "线上资源使用收入标识",
             c_orderno as "订金合同编号",
             cust_id    as "资源ID",
             zt         as "账套",
             sup_id     as "补充协议ID",
             con_type   as "合同类型",
             con_money  as "合同金额",
             deduct_money  as "代金券抵扣金额",
             signed_type as "签单类型",
             service_dept as "现服务部门",
             shop_id as "店铺ID",
             brand as "品牌"
        from cv_collect_update
       where tdate >= '2022-06-01' and nc_shop_name = '{sname}' 
       order by tdate, paytime, type, orderno
    '''
  },
  'vip_payment': {
    'memo': '当月一对一到款',
    'sql': '''
      select '序号', '日期', '姓名', '合计', '拉卡拉', '数字王府井', '银行到款', '缘分期到款'
      union all
      select ifnull(t.rn, '合计') 序号,
             if(t.rn is null, '', date_format(t.paytime, '%Y-%m-%d')) 日期,
             if(t.rn is null, '', t.uname) 姓名,
             sum(t.amount) 合计,
             sum(t.lakala) 拉卡拉,
             sum(t.wangfujing) 数字王府井,
             sum(t.remit_amt) 银行到款,
             sum(t.loan_amt) 缘分期到款
        from (select row_number() over(order by paytime) rn,
                     paytime,
                     uname,
                     lakala + wangfujing + remit_amt + loan_amt as amount,
                     lakala,
                     wangfujing,
                     remit_amt,
                     loan_amt
                from cv_collect_result
               where tdate >= '{v_date}'
                 and tdate < date_add('{v_date}', interval 1 month)
                 and nc_shop_name = '{sname}'
                 and orderno not like '%试机%'
                 and balance >= 1
                 and itype <> -4) t
       group by t.rn with rollup
    '''
  },
  'qg_payment': {
    'memo': '当月情感到款',
    'sql': '''
      select '序号', '日期', '姓名', '合计', '拉卡拉', '数字王府井', '银行到款', '缘分期到款'
      union all
      select ifnull(t.rn, '合计') 序号,
             if(t.rn is null, '', date_format(t.paytime, '%Y-%m-%d')) 日期,
             if(t.rn is null, '', t.uname) 姓名,
             sum(t.amount) 合计,
             sum(t.lakala) 拉卡拉,
             sum(t.wangfujing) 数字王府井,
             sum(t.remit_amt) 银行到款,
             sum(t.loan_amt) 缘分期到款
        from (select row_number() over(order by paytime) rn,
                     paytime,
                     uname,
                     lakala + wangfujing + remit_amt + loan_amt as amount,
                     lakala,
                     wangfujing,
                     remit_amt,
                     loan_amt
                from cv_collect_result
               where tdate >= '{v_date}'
                 and tdate < date_add('{v_date}', interval 1 month)
                 and nc_shop_name = '{sname}'
                 and orderno not like '%试机%'
                 and balance >= 1
                 and itype = -4) t
       group by t.rn with rollup
    '''
  },
  'err_update': {
    'memo': '更新支付信息异常表审核通过时间与客户ID',
    'sql': '''
      update cv_collect_err a
         set a.audit_time =
             (select distinct audit_time
                from cv_order_#tab_name b
               where a.orderno = b.orderno),
             a.cust_id   =
             (select distinct cust_id
                from cv_order_#tab_name b
               where a.orderno = b.orderno)
       where tdate = '#v_date'
         and exists
       (select 1 from cv_order_#tab_name b where a.orderno = b.orderno)
    '''
  },
  'err_detail': {
    'memo': '支付信息异常明细',
    'sql': '''
      select date_format(tdate, '%Y-%m-%d') as "归属月",
             date_format(txntime, '%Y-%m-%d') as "交易时间",
             sname as "店铺简称",
             amount as "交易金额",
             terminalid as "终端编号",
             rrn as "交易参考号",
             channel_type as "交易第三方主体（王府井/拉卡拉）",
             shop_name as "门店名称",
             terminalname as "terminalname"
        from cv_pos_detail
       where tdate = '{v_date}'
         and not (txntime >= date_sub(tdate, interval 1 day) and
              txntime < date_add(tdate, interval 1 month))
    '''
  },
  'pos_err': {
    'memo': 'pos支付异常',
    'sql': '''
      select date_format(tdate, '%Y-%m-%d') as "归属月",
             date_format(txntime, '%Y-%m-%d') as "交易时间",
             sname as "店铺简称",
             amount as "交易金额",
             terminalid as "终端编号",
             rrn as "交易参考号",
             channel_type as "交易第三方主体（王府井/拉卡拉）",
             shop_name as "门店名称",
             terminalname as "terminalname"
        from cv_pos_detail
       where tdate = '{v_date}' 
         and not( txntime >= date_sub(tdate, interval 1 day) and txntime < date_add(tdate, interval 1 month))
    '''
  },
}


def exp_duizh_result():
  if os.path.exists('/u01/vip48dian/lit/fentan/duizh/excel/'):
    # 删除文件夹
    shutil.rmtree('/u01/vip48dian/lit/fentan/duizh/excel/')
    # 创建文件夹
    os.mkdir('/u01/vip48dian/lit/fentan/duizh/excel/')
  else:
    os.mkdir('/u01/vip48dian/lit/fentan/duizh/excel/')
  
  v_sql = '''
  select distinct nc_shop_name 
    from cv_collect_result
   where tdate >= '2022-06-01' 
  '''
  snames = rordb.select(v_sql)
  
  for i in snames:
    v_sname = i[0]
    # 汇总新增表明细
    df1 = pd.read_sql(sql_dict['collect_add']['sql'].format(sname=v_sname), rordb)
    df1 = df1.fillna('')
    # pos到款拉卡拉明细
    df2 = pd.read_sql(sql_dict['lakala_detail']['sql'].format(sname=v_sname), rordb)
    df2 = df2.fillna('')
    # pos到款王府井明细
    df3= pd.read_sql(sql_dict['wangfujing_detail']['sql'].format(sname=v_sname), rordb)
    df3= df3.fillna('')
    # 汇总调账表明细
    df11= pd.read_sql(sql_dict['collect_update']['sql'].format(sname=v_sname), rordb)
    df11= df11.fillna('')
    # 一对一到款
    df12= pd.read_sql(sql_dict['vip_payment']['sql'].format(v_date=v_date, sname=v_sname), rordb)
    df12= df12.fillna('')
    # 情感到款
    df13= pd.read_sql(sql_dict['qg_payment']['sql'].format(v_date=v_date, sname=v_sname), rordb)
    df13= df13.fillna('')

    duizh_xlsx = '/u01/vip48dian/lit/fentan/duizh/excel/{v_sname}_{lbday_m}.xlsx'.format(v_sname=v_sname, lbday_m = lbday_m)
  
    # 生成excel
    writer = pd.ExcelWriter(duizh_xlsx)
    df1.to_excel(writer, sheet_name='汇总', index=False, startrow=0, startcol=0)
    df2.to_excel(writer, sheet_name='拉卡拉', index=False, startrow=0, startcol=0)
    df3.to_excel(writer, sheet_name='王府井', index=False, startrow=0, startcol=0)
    df11.to_excel(writer, sheet_name='调账', index=False, startrow=0, startcol=0)
    df12.to_excel(writer, sheet_name='一对一到款', index=False, startrow=0, startcol=0)
    df13.to_excel(writer, sheet_name='情感到款', index=False, startrow=0, startcol=0)
    workbook = writer.book
    worksheets = writer.sheets
    worksheet1 = worksheets['汇总']
    worksheet2 = worksheets['拉卡拉']
    worksheet3 = worksheets['王府井']
    worksheet4 = worksheets['调账']
    worksheet5 = worksheets['一对一到款']
    worksheet6 = worksheets['情感到款']
    
    header_format = workbook.add_format(__add_format__('header'))
    body_format = workbook.add_format(__add_format__('body'))
    header_ch_format = workbook.add_format(__add_format__('header_ch'))
    # 设置表头格式
    header_dict = __tup_to_dict__(((0, 1, header_ch_format), (2, 3, header_format), (4, 10, header_ch_format), (11, 21, header_format), (22, 22, header_ch_format), (23, 24, header_format), (25, 32, header_ch_format)))
    __write_sheet__(df1, worksheet1, 0, 0, header_format, body_format, 0, '', header_dict)
    __write_sheet__(df2, worksheet2, 0, 0, header_format, body_format, 0, '', {})
    __write_sheet__(df3, worksheet3, 0, 0, header_format, body_format, 0, '', header_dict)
    __write_sheet__(df11, worksheet4, 0, 0, header_format, body_format, 0, '', {})
    __write_sheet__(df12, worksheet5, 1, 0, header_format, body_format, 0, '', {})
    __write_sheet__(df13, worksheet6, 1, 0, header_format, body_format, 0, '', {})
    worksheet5.merge_range('A1:H1',
                           f"{v_sname.split('-')[1]}一对一确认到款表({str(int(v_date[5:7]))}月)",
                           header_ch_format)
    worksheet6.merge_range('A1:H1',
                           f"{v_sname.split('-')[1]}情感到款表({str(int(v_date[5:7]))}月)",
                           header_ch_format)
    for sheet_num in (worksheet5, worksheet6):
      sheet_num.set_column('A: A', 20)
      sheet_num.merge_range('A2:A3', df12.iloc[[0], [0]].values[0][0], header_format)
      sheet_num.merge_range('B2:B3', df12.iloc[[0], [1]].values[0][0], header_format)
      sheet_num.merge_range('C2:C3', df12.iloc[[0], [2]].values[0][0], header_format)
      sheet_num.merge_range('D2:D3', df12.iloc[[0], [3]].values[0][0], header_format)
      sheet_num.merge_range('E2:F2', 'pos到款', header_format)
      sheet_num.merge_range('G2:G3', df12.iloc[[0], [6]].values[0][0], header_format)
      sheet_num.merge_range('H2:H3', df12.iloc[[0], [7]].values[0][0], header_format)
      sheet_num.conditional_format('E3:F3', {'type': 'blanks', 'format': header_format})

    writer.save()
    # 使用openpyxl设置日期格式
    wb = openpyxl.load_workbook(duizh_xlsx)
    for sheet in wb.sheetnames:
      wc = wb[sheet]
      for i in range(wc.max_row):
        if sheet in ('汇总', '拉卡拉', '王府井', '调账'):
          wc.cell(i+1, 1).number_format='yyyy-mm-dd' 
          if sheet in ('汇总', '调账'):
            wc.cell(i+1, 7).number_format='yyyy-mm-dd'
            wc.cell(i+1, 10).number_format='yyyy-mm-dd'
          if sheet in ('拉卡拉', '王府井'):
            wc.cell(i+1, 3).number_format='yyyy-mm-dd'
        elif sheet in ('一对一到款', '情感到款'):
          wc.cell(i+3, 2).number_format='yyyy-mm-dd'
    wb.save(duizh_xlsx)

  # jy_db.close()
  # bh_db.close()
  
  zipname = u'/u01/vip48dian/lit/fentan/duizh/res/duizhang_result_{lbday_m}.zip'.format(lbday_m=lbday_m)
  if os.path.exists(zipname):
    os.remove(zipname)
  # 压缩excel文件
  logdir = '/u01/vip48dian/lit/fentan/duizh/excel/'
  #创建zip对象，
  fzip = zipfile.ZipFile(zipname, 'w', zipfile.ZIP_DEFLATED)
  #遍历要压缩目录
  flist = os.listdir(logdir)
  #获取压缩目录名称
  basename = os.path.basename(logdir)
  for name in flist:
    if lbday_m in name:
      fpath = os.path.join(logdir, name)
      arcname = os.path.join(basename, name)
      #写入要压缩文件，并添加归档文件名称
      fzip.write(fpath, arcname=arcname)
  #关闭
  fzip.close()
  
  tab_name = v_date.replace('-', '')[0: 6]
  # 更新支付信息异常表审核通过时间与客户ID
  rordb.update(sql_dict['err_update']['sql'], tab_name=tab_name, v_date=v_date)
  # 支付信息异常明细
  v_sql3 = sql_dict['err_detail']['sql'].format(v_date=v_date)
  df4 = pd.read_sql(v_sql3, rordb)
  df4 = df4.fillna('')
  duizh_err_xlsx = '/u01/vip48dian/lit/fentan/duizh/res/duizhang_payerr_{lbday_m}.xlsx'.format(lbday_m = lbday_m)
  # 生成excel
  writer = pd.ExcelWriter(duizh_err_xlsx)
  df4.to_excel(writer, sheet_name='支付异常信息', index=False, startrow=0, startcol=0)
  workbook = writer.book
  worksheets = writer.sheets
  worksheet1 = worksheets['支付异常信息']
  header_format = workbook.add_format(__add_format__('header'))
  body_format = workbook.add_format(__add_format__('body'))
  header_ch_format = workbook.add_format(__add_format__('header_ch'))
  __write_sheet__(df4, worksheet1, 0, 0, header_format, body_format, 0, '', {})
  writer.save()
  
  pos_err = 0
  # pos支付异常
  pos_data = rordb.select(sql_dict['err_update']['sql'].format(v_date=v_date))
  if pos_data and pos_data[0][0]:
    df5 = pd.read_sql(v_sql3, rordb)
    df5 = df5.fillna('')
    pos_err_xlsx = '/u01/vip48dian/lit/fentan/duizh/res/duizhang_poserr_{lbday_m}.xlsx'.format(lbday_m = lbday_m)
    # 生成excel
    writer = pd.ExcelWriter(pos_err_xlsx)
    df5.to_excel(writer, sheet_name='pos支付录入异常信息', index=False, startrow=0, startcol=0)
    workbook = writer.book
    worksheets = writer.sheets
    worksheet1 = worksheets['pos支付录入异常信息']
    header_format = workbook.add_format(__add_format__('header'))
    body_format = workbook.add_format(__add_format__('body'))
    header_ch_format = workbook.add_format(__add_format__('header_ch'))
    __write_sheet__(df5, worksheet1, 0, 0, header_format, body_format, 0, '', {})
    writer.save()
    pos_err = 1

  # 导出excel
  if pos_err == 1:
    curl_txt = '''curl --form "duizhang_result_{lbday_m}.zip=@{zipname}" --form "duizhang_payerr_{lbday_m}.xlsx=@{duizh_err_xlsx}"  --form "duizhang_poserr_{lbday_m}.xlsx=@{pos_err_xlsx}" --form toMails={pos_name} --form subject='对账结果_{lbday_m}'  --form body='对账结果_{lbday_m}' 'http://jf-houtai.jiayuan.idc:8080/alarm/SendMail2.jsp?from=jf_jk@jiayuan.com&pass=jf_jk)(*' 2>/dev/null'''
    curlf = curl_txt.format(lbday_m=lbday_m, zipname=zipname, pos_name=pos_name, duizh_err_xlsx=duizh_err_xlsx, pos_err_xlsx=pos_err_xlsx)
  else:
    curl_txt = '''curl --form "duizhang_result_{lbday_m}.zip=@{zipname}" --form "duizhang_payerr_{lbday_m}.xlsx=@{duizh_err_xlsx}" --form toMails={pos_name} --form subject='对账结果_{lbday_m}'  --form body='对账结果_{lbday_m}' 'http://jf-houtai.jiayuan.idc:8080/alarm/SendMail2.jsp?from=jf_jk@jiayuan.com&pass=jf_jk)(*' 2>/dev/null'''
    curlf = curl_txt.format(lbday_m=lbday_m, zipname=zipname, pos_name=pos_name, duizh_err_xlsx=duizh_err_xlsx)
  print(curlf)
  os.popen(curlf)

try:
  exp_duizh_result()
except Exception as e:
  traceback.print_exc()
  print(e)
finally:
  rordb.close()

e_time = time.time()
print('执行时间%0.3fs' % (e_time - b_time))
